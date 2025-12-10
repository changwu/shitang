import os
import argparse
import csv
import json
import re
import logging
import traceback
from datetime import datetime, date
from typing import List, Dict, Any, Tuple

from sqlalchemy import create_engine, MetaData, Table, Column, func, select
from sqlalchemy import Integer, Float, Boolean, Date, DateTime, Text, BigInteger, Numeric
from decimal import Decimal
from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def sanitize_identifier(name: str) -> str:
    s = re.sub(r"[^0-9a-zA-Z_]+", "_", name.strip())
    s = s.lower()
    if not s or not re.match(r"^[a-zA-Z_]", s):
        s = f"col_{s}" if s else "col"
    return s


def sanitize_table(name: str) -> str:
    s = sanitize_identifier(name)
    return s


def to_bool(v: str) -> Tuple[bool, bool]:
    lv = v.strip().lower()
    if lv in {"true", "t", "1", "yes", "y"}:
        return True, True
    if lv in {"false", "f", "0", "no", "n"}:
        return False, True
    return False, False


def try_int(v: str) -> Tuple[int, bool]:
    if re.fullmatch(r"[+-]?\d+", v.strip() or ""):
        try:
            return int(v), True
        except Exception:
            return 0, False
    return 0, False


def try_float(v: str) -> Tuple[float, bool]:
    try:
        if v.strip() == "":
            return 0.0, False
        return float(v), True
    except Exception:
        return 0.0, False


def try_date(v: str) -> Tuple[Any, bool, bool]:
    s = v.strip()
    if not s:
        return None, False, False
    try:
        if "T" in s or "+" in s or s.endswith("Z"):
            s2 = s.replace("Z", "+00:00")
            return datetime.fromisoformat(s2), True, True
        return date.fromisoformat(s), True, False
    except Exception:
        return None, False, False


def infer_sqlalchemy_type(samples: List[Any]) -> Any:
    ints = 0
    bigints = 0
    floats = 0
    bools = 0
    dates = 0
    datetimes = 0
    total = 0
    for sv in samples:
        if sv is None:
            continue
        total += 1
        if isinstance(sv, bool):
            bools += 1
            continue
        if isinstance(sv, int):
            ints += 1
            if abs(sv) > 2_147_483_647:
                bigints += 1
            continue
        if isinstance(sv, float):
            floats += 1
            continue
        if isinstance(sv, datetime):
            datetimes += 1
            continue
        if isinstance(sv, date):
            dates += 1
            continue
    if total == 0:
        return Text
    if bools == total:
        return Boolean
    if datetimes == total:
        return DateTime
    if dates == total:
        return Date
    if floats == total:
        return Float
    if ints == total:
        return BigInteger if bigints > 0 else Integer
    return Text


def normalize_value(v: Any, typ: Any) -> Any:
    if v is None:
        return None
    if typ is DateTime and isinstance(v, datetime):
        return v
    if typ is Date and isinstance(v, date) and not isinstance(v, datetime):
        return v
    if typ is Boolean:
        b, ok = to_bool(str(v))
        return b if ok else None
    if typ is Integer:
        i, ok = try_int(str(v))
        return i if ok else None
    if typ is BigInteger:
        i, ok = try_int(str(v))
        return i if ok else None
    if typ is Float:
        f, ok = try_float(str(v))
        return f if ok else None
    if typ is DateTime:
        dt, ok, is_dt = try_date(str(v))
        return dt if ok and is_dt else None
    if typ is Date:
        d, ok, is_dt = try_date(str(v))
        return d if ok and not is_dt else None
    if typ is Numeric:
        try:
            if v is None or (isinstance(v, str) and v.strip() == ""):
                return None
            return Decimal(str(v))
        except Exception:
            return None
    return v if v != "" else None


def read_csv(path: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({k: (None if (v is None or v == "" or str(v).lower() == "null") else v) for k, v in r.items()})
    return rows


def read_json(path: str) -> List[Dict[str, Any]]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    data.append(obj)
                except Exception:
                    pass
    if isinstance(data, dict):
        for k in ("data", "items", "records"):
            if k in data and isinstance(data[k], list):
                data = data[k]
                break
        if isinstance(data, dict):
            data = [data]
    rows: List[Dict[str, Any]] = []
    for item in data if isinstance(data, list) else []:
        if isinstance(item, dict):
            flat: Dict[str, Any] = {}
            for k, v in item.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = v
            rows.append(flat)
    return rows


def read_xlsx(path: str, header_row: int = 1) -> List[Tuple[str, List[Dict[str, Any]], List[str]]]:
    if not OPENPYXL_AVAILABLE:
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    results: List[Tuple[str, List[Dict[str, Any]], List[str]]] = []
    for ws in wb.worksheets:
        sheet_name = ws.title
        headers: List[str] = []
        # read header row (1-based index)
        hdr_cells = ws[header_row]
        for idx, cell in enumerate(hdr_cells, start=1):
            raw = str(cell.value) if cell.value is not None else f"col_{idx}"
            headers.append(raw)
        rows: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row + 1):
            record: Dict[str, Any] = {}
            for i, cell in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i+1}"
                val = cell.value
                if val is None:
                    record[key] = None
                    continue
                try:
                    if getattr(cell, "is_date", False):
                        if isinstance(val, (int, float)):
                            record[key] = from_excel(val)
                        else:
                            record[key] = val
                    else:
                        record[key] = val
                except Exception:
                    record[key] = val
            rows.append(record)
        results.append((sheet_name, rows, headers))
    return results


# 新的表结构定义 - 使用英文字段名
TABLES = {
    "canteen_records": {
        "columns": {
            "id": Integer,
            "record_date": DateTime(timezone=True),
            "name": Text,
            "type": Text,
        },
        "excel_mappings": {
            "消费时间": "record_date",
            "姓名": "name", 
            "餐别": "type"
        }
    },
    "vehicle_records": {
        "columns": {
            "id": Integer,
            "record_date": DateTime(timezone=True),
            "name": Text,
            "type": Text,
        },
        "excel_mappings": {
            "打卡时间": "record_date",
            "姓名": "name",
            "打卡类型": "type"
        }
    },
    "door_records": {
        "columns": {
            "id": Integer,
            "record_date": DateTime(timezone=True),
            "name": Text,
            "type": Text,
        },
        "excel_mappings": {
            "事件时间": "record_date",
            "人员姓名": "name",
            "控制器": "type"
        }
    },
}


def ensure_table(metadata: MetaData, table_name: str, columns: Dict[str, Any]) -> Table:
    try:
        tbl = Table(table_name, metadata, autoload_with=metadata.bind)
        return tbl
    except Exception:
        pass
    cols = [Column(cn, ct, quote=True) for cn, ct in columns.items()]
    cols.append(Column("created_at", DateTime(timezone=True), server_default="now()", quote=True))
    tbl = Table(table_name, metadata, *cols)
    metadata.create_all(metadata.bind, tables=[tbl])
    return tbl


def process_excel_for_table(engine, metadata: MetaData, path: str, excel_header_row: int, target_table: str, verbose: bool) -> None:
    xresults = read_xlsx(path, header_row=excel_header_row)
    if not xresults:
        return
    
    schema = TABLES[target_table]["columns"]
    excel_mappings = TABLES[target_table]["excel_mappings"]
    tbl = ensure_table(metadata, target_table, schema)
    
    def _canon(h: Any) -> str:
        s = str(h) if h is not None else ""
        s = s.strip()
        s = s.replace("（", "(").replace("）", ")")
        s = re.sub(r"\s+", "", s)
        return s
    
    for sheet_name, srows, orig_headers in xresults:
        if verbose:
            logging.info(f"[import] file={os.path.basename(path)} sheet={sheet_name} rows={len(srows)} -> table={target_table}")
        
        header_map: Dict[str, str] = {}
        orig_map: Dict[str, str] = {_canon(h): h for h in orig_headers}
        for source_header, target_field in excel_mappings.items():
            canon_source = _canon(source_header)
            if canon_source in orig_map:
                header_map[target_field] = orig_map[canon_source]
            else:
                header_map[target_field] = source_header
        
        ins_rows = []
        for r in srows:
            ir: Dict[str, Any] = {}
            # 手动设置id为自增
            ir["id"] = None  # 数据库会自动处理自增
            
            for source_header, target_field in excel_mappings.items():
                field_type = schema[target_field]
                hk = header_map.get(target_field, source_header)
                raw = r.get(hk)
                ir[target_field] = normalize_value(raw, field_type)
            
            ins_rows.append(ir)
        
        if ins_rows:
            if verbose:
                logging.info(f"Inserting {len(ins_rows)} rows into {target_table}")
            with engine.begin() as conn:
                # 插入时排除id字段，让数据库自动处理
                insert_data = []
                for row in ins_rows:
                    filtered_row = {k: v for k, v in row.items() if k != "id"}
                    insert_data.append(filtered_row)
                
                if insert_data:
                    conn.execute(tbl.insert(), insert_data)
            
            if verbose:
                logging.info(f"Successfully inserted {len(ins_rows)} rows into {target_table}")


def determine_table_by_filename(filename: str) -> str:
    """
    根据文件名确定目标表
    文件名含有"consumelog"的excel文件对应食堂消费记录表"canteen_records"
    文件名含有"打卡明细数据"的excel文件对应车辆打卡记录表"vehicle_records" 
    文件名含有"dooreventinfo"的excel文件对应人员打卡记录表"door_records"
    """
    fname = filename.lower()
    
    if "consumelog" in fname:
        return "canteen_records"
    elif "打卡明细数据" in fname:
        return "vehicle_records"
    elif "dooreventinfo" in fname:
        return "door_records"
    else:
        return None


def process_file(engine, metadata: MetaData, path: str, excel_header_row: int = 1, verbose: bool = False, only_tables: List[str] = None) -> None:
    base = os.path.basename(path)
    name, ext = os.path.splitext(base)
    if ext.lower() not in {".xlsx", ".xlsm"}:
        return
    
    # 根据文件名确定目标表
    target_table = determine_table_by_filename(name)
    
    if not target_table:
        if verbose:
            logging.warning(f"无法根据文件名确定目标表: {path}")
            logging.warning(f"文件名需要包含: 'consumelog', '打卡明细数据', 或 'dooreventinfo'")
        return
    
    # 检查是否只允许特定表
    if only_tables and target_table not in only_tables:
        if verbose:
            logging.info(f"跳过 {target_table} 对于文件 {path} (不在 --table 过滤器中)")
        return
    
    if verbose:
        logging.info(f"处理文件 {path} -> 表 {target_table}")
    
    process_excel_for_table(engine, metadata, path, excel_header_row, target_table, verbose)


def main():
    load_dotenv()
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", default=os.environ.get("DATA_DIR", "./data"))
    parser.add_argument("--import-dir", default=os.environ.get("IMPORT_DIR", "import/"))
    parser.add_argument("--db-url", default=os.environ.get("DB_URL", "postgresql+psycopg2://postgres:postgres@localhost:5432/shitang"))
    parser.add_argument("--excel-header-row", type=int, default=int(os.environ.get("EXCEL_HEADER_ROW", "1")))
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--table", choices=["canteen_records", "vehicle_records", "door_records"], nargs="*")
    args = parser.parse_args()
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    engine = create_engine(args.db_url, future=True)
    metadata = MetaData()
    metadata.bind = engine
    
    # Log import start
    logging.info("开始导入过程...")
    
    # Get initial counts for each table
    initial_counts = {}
    for table_name in ["canteen_records", "vehicle_records", "door_records"]:
        try:
            tbl = Table(table_name, metadata, autoload_with=engine)
            with engine.connect() as conn:
                result = conn.scalar(select(func.count()).select_from(tbl))
                initial_counts[table_name] = result
                logging.info(f"初始计数 {table_name}: {initial_counts[table_name]}")
        except Exception as e:
            initial_counts[table_name] = 0
            logging.warning(f"无法获取 {table_name} 的初始计数: {e}")
    
    base_dir = os.path.join(args.data_dir, args.import_dir)
    processed_files = 0
    
    for root, _, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith((".xlsx", ".xlsm")):
                file_path = os.path.join(root, f)
                try:
                    process_file(engine, metadata, file_path, excel_header_row=args.excel_header_row, verbose=args.verbose, only_tables=args.table)
                    processed_files += 1
                    logging.info(f"成功处理: {file_path}")
                except Exception as e:
                    logging.error(f"处理 {file_path} 失败: {e}")
                    logging.error(traceback.format_exc())
    
    # Log final counts and changes
    logging.info("导入过程完成!")
    logging.info(f"总共处理文件: {processed_files}")
    
    for table_name in ["canteen_records", "vehicle_records", "door_records"]:
        try:
            tbl = Table(table_name, metadata, autoload_with=engine)
            with engine.connect() as conn:
                result = conn.scalar(select(func.count()).select_from(tbl))
                final_count = result
                initial_count = initial_counts.get(table_name, 0)
                change = final_count - initial_count
                logging.info(f"最终计数 {table_name}: {final_count} (变化: +{change})")
        except Exception as e:
            logging.warning(f"无法获取 {table_name} 的最终计数: {e}")


if __name__ == "__main__":
    main()
